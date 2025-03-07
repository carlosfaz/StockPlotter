from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

def procesar_excel(filename):
    workbook = load_workbook(filename)
    gris_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    columnas_porcentaje = ["L", "M", "N", "O"]  # Especifica las columnas de porcentaje

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Agregar hipervínculos en la columna B
        for row in sheet.iter_rows(min_col=2, max_col=2, min_row=2):  # Comienza desde la segunda fila
            cell = row[0]
            if cell.value:  # Si la celda tiene un valor
                ticker = str(cell.value).strip()  # Asegurarse de que el ticker esté bien formateado
                url = f"https://finance.yahoo.com/quote/{ticker}"
                # Crear el hipervínculo en la celda
                cell.hyperlink = url
                cell.value = ticker  # El texto visible es el ticker
                cell.style = 'Hyperlink'  # Establecer el estilo de hipervínculo

        # Ajuste del tamaño de las columnas basado en su contenido
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[column_letter].width = max_length + 2

        # Relleno gris en la primera fila
        for cell in sheet[1]:
            cell.fill = gris_fill

        # Ajuste de columnas de porcentaje
        for col in columnas_porcentaje:
            if col in sheet.column_dimensions:  # Verifica si la columna existe
                for cell in sheet[col][1:]:  # Excluye el encabezado
                    if isinstance(cell.value, (int, float)):
                        cell.value = cell.value / 100  # Dividir entre 100 para corregir el porcentaje
                        cell.number_format = '0.00%'  # Aplicar formato de porcentaje

    workbook.save(filename)
    print(f"Archivo excel procesado en {filename}")
